import openpyxl
import pandas as pd
import os
import sys
import numpy as np
from openpyxl import load_workbook
from Datos_ETs_CAs import datos_ets_cas
from Datos_Loc_CAs import datos_loc_cas
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter



from Datos_Loc_ETs import datos_loc_ets
from Utiles import calcular_termino_producto_escalar2, calcular_termino_producto_escalar3, evaluarCuartaRestriccion, evaluarPrimeraRestriccion, evaluarSegundaRestriccion, evaluarTerceraRestriccion, funcion_ci_ca, funcion_ci_et, funcion_co_ca, funcion_co_et, generar_matriz, generar_matriz_nueva, generarNuevoBjK, productoEscalar2Matrices, productoEscalar3Matrices, productoEscalar4Matrices
from prueba import matriz_iterada
from datetime import datetime

# Obtiene la ruta del directorio actual
ruta_actual = os.path.dirname(os.path.abspath(sys.argv[0]))

# Nombre del archivo Excel que quieres leer
nombre_archivo = 'modelo.xlsm'  # Reemplaza con el nombre de tu archivo Excel

# Une la ruta del directorio actual con el nombre del archivo Excel
ruta_archivo = os.path.join(ruta_actual, nombre_archivo)

# Lee todas las hojas del archivo Excel en un diccionario
datos_excel = pd.read_excel(ruta_archivo, sheet_name=None)
nuevo_arreglo = []
INVTSj = 0  # Costo de Inversión ET
OPTSj = 0  # Costo Operativo ET
RECj = 0  # Tasa de material recuperado et
SPj = 0  # Costo de venta de material recuperado et
RECk = 0  # Tasa de material recuperado ca
SPk = 0  # Costo de venta de material recuperado ca
INVLk = 0  # Costo de inversión de Centro Ambiental
OPLk = 0  # Costo operativo de Centro Ambiental	
cantidad_de_ks = 0 # Cantidad de Centros ambientales definidos en la celda G31
CC = 0 # Costo de transporte de camiones recolectores
TC = 0 # Costo de transporte de camiones de transferencia
Djk = [[]]
Dik = [[]]
Dij = [[]]
Vc = 0 #Capacidad en m3 de camion recolector
Ec = 0 #Eficiencia de carga en % recolector
Dc = 0 #Densidad en tn/m3 de RSU en camion recolector
Rc = 0 #Tasa de reserva de camiones recolectores
T1c = 0 # Horas de trabajo de camion recolector
T2c = 0 # Horas de preparación de camion recolector
T3c = 0 # Horas de limpieza de camion recolector
T4c = 0 # Horas de carga de camion recolector
T5c = 0 # Horas de descarga de camion recolector
SPDc = 0 # Velocidad de camion recolector
Vt = 0 #Capacidad en m3 de camion transportador
Et = 0 #Eficiencia de carga en % transportador
Dt = 0 #Densidad en tn/m3 de RSU en camion transportador
Rt = 0 #Tasa de reserva de camiones transportadores
T1t = 0 # Horas de trabajo de camion transportador
T2t = 0 # Horas de preparación de camion transportador
T3t = 0 # Horas de limpieza de camion transportador
T4t = 0 # Horas de carga de camion transportador
T5t = 0 # Horas de descarga de camion transportador
SPDt = 0 # Velocidad de camion transportador
Uc = 0 # Factor U para recolectores
KPLc = 0 # distancia recorrida en un litro de camiones recolectores km/l
Ut = 0 # Factor U para transportadores
KPLt = 0 # distancia recorrida en un litro de camiones transportadores km/l
n_iter = 0 #Numero de iteraciones máximas a realizar por el modelo
rango_et = 0 #% de Generación como rango de ET
rango_ca = 0 #% de Generación como rango de CA
promedio_gen_inicial = 0 # Punto de quiebre entre ET de 1 piso y ET de 2 pisos
vida_util_et = 0 #Vida util ET
vida_util_ca = 0 #Vida util CA
tasainteres = 0 #Tasa de interes anual
texto_region = 0
texto_i1 = 0
texto_i2 = 0
texto_i3 = 0
texto_i4 = 0
texto_i5 = 0
texto_i10 = 0
texto_i7 = 0
hoja_INPUTS_Generales = datos_excel.get('INPUTS_Generales')
hoja_Datos_Loc_ETs = datos_excel.get('Datos_Loc_ETs')
hoja_Datos_Loc_CAs = datos_excel.get('Datos_Loc_CAs')
hoja_Datos_ETs_CAs = datos_excel.get('Datos_ETs_CAs')
hoja_Gen_Cap = datos_excel.get('Gen-Cap')


if hoja_INPUTS_Generales is not None:
    INVTSj = hoja_INPUTS_Generales.iloc[20, 6]
    OPTSj = hoja_INPUTS_Generales.iloc[21, 6]
    RECj = hoja_INPUTS_Generales.iloc[22, 6]
    SPj = hoja_INPUTS_Generales.iloc[23, 6]
    RECk = hoja_INPUTS_Generales.iloc[27, 6]
    SPk = hoja_INPUTS_Generales.iloc[28, 6]
    INVLk = hoja_INPUTS_Generales.iloc[25, 6]
    OPLk = hoja_INPUTS_Generales.iloc[26, 6]
    cantidad_de_ks = hoja_INPUTS_Generales.iloc[33,6]
    CC = hoja_INPUTS_Generales.iloc[14,5]
    TC = hoja_INPUTS_Generales.iloc[14,6]
    Vc = hoja_INPUTS_Generales.iloc[8,5]
    Ec = hoja_INPUTS_Generales.iloc[9,5]
    Dc = hoja_INPUTS_Generales.iloc[12,5]
    Rc = hoja_INPUTS_Generales.iloc[13,5]
    T1c = hoja_INPUTS_Generales.iloc[2,5]
    T2c = (hoja_INPUTS_Generales.iloc[3,5])/60
    T3c = (hoja_INPUTS_Generales.iloc[4,5])/60
    T4c = hoja_INPUTS_Generales.iloc[5,5]
    T5c = (hoja_INPUTS_Generales.iloc[6,5])/60
    SPDc = hoja_INPUTS_Generales.iloc[7,5]
    Vt = hoja_INPUTS_Generales.iloc[8,6]
    Et = hoja_INPUTS_Generales.iloc[9,6]
    Dt = hoja_INPUTS_Generales.iloc[12,6]
    Rt = hoja_INPUTS_Generales.iloc[13,6]
    T1t = hoja_INPUTS_Generales.iloc[2,6]
    T2t = (hoja_INPUTS_Generales.iloc[3,6])/60
    T3t = (hoja_INPUTS_Generales.iloc[4,6])/60
    T4t = hoja_INPUTS_Generales.iloc[5,6]
    T5t = (hoja_INPUTS_Generales.iloc[6,6])/60
    SPDt = hoja_INPUTS_Generales.iloc[7,6]
    Uc = hoja_INPUTS_Generales.iloc[17,5]
    KPLc = hoja_INPUTS_Generales.iloc[15,5]
    Ut = hoja_INPUTS_Generales.iloc[17,6]
    KPLt = hoja_INPUTS_Generales.iloc[15,6]
    n_iter = hoja_INPUTS_Generales.iloc[37,6]
    rango_et = hoja_INPUTS_Generales.iloc[38,6]
    rango_ca = hoja_INPUTS_Generales.iloc[39,6]
    promedio_gen_inicial = hoja_INPUTS_Generales.iloc[40,6]
    vida_util_et = hoja_INPUTS_Generales.iloc[24, 6]
    vida_util_ca = hoja_INPUTS_Generales.iloc[29, 6]
    tasainteres = hoja_INPUTS_Generales.iloc[41, 6]
    texto_region = hoja_INPUTS_Generales.iloc[31, 10]
    texto_i1 = hoja_INPUTS_Generales.iloc[33, 10] 
    texto_i2 = hoja_INPUTS_Generales.iloc[34, 10]
    texto_i3 = hoja_INPUTS_Generales.iloc[35, 10]
    texto_i4 = hoja_INPUTS_Generales.iloc[36, 10]
    texto_i5 = hoja_INPUTS_Generales.iloc[37, 10]
    texto_i6 = hoja_INPUTS_Generales.iloc[38, 10]
    texto_i7 = hoja_INPUTS_Generales.iloc[39, 10]
    

if hoja_Datos_Loc_ETs is not None:
    Dij = datos_loc_ets(hoja_Datos_Loc_ETs)

if hoja_Datos_Loc_CAs is not None:
    Dik = datos_loc_cas(hoja_Datos_Loc_CAs)

if hoja_Datos_ETs_CAs is not None:
    Djk = datos_ets_cas(hoja_Datos_ETs_CAs)

if hoja_Gen_Cap is not None:
    data = pd.DataFrame(hoja_Gen_Cap)
    values_list = data.values.flatten().tolist()
    numeric_arrays = []
    current_numeric_array = []
    for val in values_list:
        if val == '[t/d]':
            if current_numeric_array:
                numeric_arrays.append([x for x in current_numeric_array if not np.isnan(x)])
                current_numeric_array = []
        elif isinstance(val, (int, float)):
            current_numeric_array.append(val)

    # Añadir el último array numérico si existe después de la última aparición de '[t/d]'
    if current_numeric_array:
        numeric_arrays.append([x for x in current_numeric_array if not np.isnan(x)])

    # Obtén la longitud del array numeric_values
    n = len(numeric_arrays[0])
    # Genera una matriz cuadrada de tamaño nxn inicializada con ceros
    B=Dij
    A=np.array(numeric_arrays[0])
    Aij = np.tile(A.reshape((n, 1)), B.shape[1])
    Aij= Aij.T
    
    TSj = generar_matriz(numeric_arrays[1])
    Ik = generar_matriz(numeric_arrays[2])
    m = len(numeric_arrays[2])
    Cik = np.zeros((n,m))
    Cik = np.tile(numeric_arrays[0], (m,1))
    
n = len(Dij)
m = n
matrices_iteradas_Xij = matriz_iterada(n, m)
n = len(Dik)
m = n
matrices_iteradas_Zik = matriz_iterada(n, m)
n = len(Djk)
m = n
matrices_iteradas_Yjk = matriz_iterada(n, m)


filas, columnas = Dij.shape
matriz_optimaXij = np.ones((filas, columnas))
filas, columnas = Dik.shape
matriz_optimaZik = np.ones((filas, columnas))
filas, columnas = Djk.shape
matriz_optimaYjk = np.ones((filas, columnas))


vector_termino_primero=[]
valor_suma_minimo=999999999999999999999999999
hora_actual = datetime.now()
print("La hora actual es:", hora_actual)


######################################Prueba de cambio de valor#########################
valor_inicial_de_corrida_et = promedio_gen_inicial

arreglo_terminos_optimos = []

for indice_x in range(n_iter -1):

    matriz_transpuesta = list(map(list, zip(*Aij)))
    matriz_et = (np.array(matriz_transpuesta) * np.array(matriz_optimaXij))
    # Suma de cada columna
    sumas_columnas_et = [sum(col) for col in zip(*matriz_et)]
    sumas_no_nulas = [suma for suma in sumas_columnas_et if suma != 0]
    # Calcula el promedio de las sumas no nulas
    promedio_et = sum(sumas_no_nulas) / len(sumas_no_nulas) if sumas_no_nulas else 0
    
   
    # Nombre del archivo Excel
    # escenario = hoja_INPUTS_Generales['A2'].value
    archivo_federico = openpyxl.load_workbook('modelo.xlsm')
    # Seleccionar la hoja del archivo "Federico"
    hoja_federico = archivo_federico['INPUTS_Generales']
 
    condicion_PS_ET = hoja_federico['K24'].value
    sep_origen = (hoja_federico['K25'].value)/100
    condicion_PS_CA = hoja_federico['K29'].value

    
       
    FVP_et = (((1+tasainteres)**vida_util_et)-1)/(tasainteres*((1+tasainteres)**vida_util_et))
    FVP_ca = (((1+tasainteres)**vida_util_ca)-1)/(tasainteres*((1+tasainteres)**vida_util_ca))
    
       
  
         
    if condicion_PS_ET == "SI":
        PS_en_ET = 1
    else:
        PS_en_ET = 0
        
    if condicion_PS_CA == "SI":
        PS_en_CA = 1
    else:
        PS_en_CA = 0
       

    matriz_transpuesta_cik = list(map(list, zip(*Cik)))
    v=(np.array(matriz_optimaZik) * np.array(matriz_transpuesta_cik))
    
   
    
    matriz_transpuesta_ppp = list(map(list, zip(*Aij)))
    A = np.array(matriz_optimaXij)
    B = np.array(matriz_transpuesta_ppp)
    productos = productoEscalar2Matrices(A,B)
    vector_resultante = np.array(productos)
    vector_termino_primero = vector_resultante
    Bjk = (np.tile(vector_resultante, (cantidad_de_ks, 1)).T) * (1- RECj*PS_en_ET)
    C = np.array(matriz_optimaYjk)
    D = np.array(Bjk)
    y = C*D
     
   
    matriz_ca = (np.array(v) + np.array(y))


    # Calcular el promedio de las sumas excluyendo los valores iguales a cero
    sumas_columnas_ca = [sum(col) for col in zip(*matriz_ca)]
    sumas_no_nulas = [suma for suma in sumas_columnas_ca if suma != 0]
    # Calcula el promedio de las sumas no nulas
    promedio_ca = sum(sumas_no_nulas) / len(sumas_no_nulas) if sumas_no_nulas else 0
      

    if (valor_inicial_de_corrida_et>promedio_et*(1+rango_et/100) or valor_inicial_de_corrida_et<promedio_et*(1-rango_et/100)):
        valor_inicial_de_corrida_et = promedio_et
        
        if promedio_ca == 0:
            INVLk = 0
            OPLk = 0
        else:
            INVLk = (funcion_ci_ca(promedio_ca, vida_util_ca))/FVP_ca
            OPLk = funcion_co_ca(promedio_ca)
                     
        if promedio_et == 0:
            INVTSj = 0
            OPTSj = 0
        else:
            INVTSj = (funcion_ci_et(promedio_et,condicion_PS_ET, vida_util_et, sep_origen))/FVP_et
            OPTSj = funcion_co_et(promedio_et, condicion_PS_ET, sep_origen)
        
    for Xij in matrices_iteradas_Xij:
            for Zik in matrices_iteradas_Zik:
                for Yjk in matrices_iteradas_Yjk:
                    print ("XIJJJJJJJJJJJJJ", np.array(Xij))
                    print ("Zikkkkkkkkkkkkk", np.array(Zik))
                    if(evaluarPrimeraRestriccion(np.array(Xij),np.array(Zik))):
                        if(evaluarSegundaRestriccion (Xij,Yjk)):
                            matriz_transpuesta = list(map(list, zip(*Aij)))
                            A = np.array(Xij)
                            B = np.array(matriz_transpuesta)
                            C = np.array(TSj)
                            D = np.array(Xij) * np.array(matriz_transpuesta)
                            E = generar_matriz_nueva(numeric_arrays[1])
                            if(evaluarTerceraRestriccion(D,E)):
                                    # Verificar si las matrices tienen la misma forma
                                if A.shape != B.shape:
                                    print("Las matrices no tienen la misma forma.")
                                # else:
                                #     productos = productoEscalar2Matrices(A,B)
                                #     resultado_final = sum(productos)
                                productos = productoEscalar2Matrices(A,B)
                                vector_resultante = np.array(productos)
                                vector_termino_primero = vector_resultante
                                resultado_multiplicacion = np.dot(vector_resultante, C)
                                primer_termino = resultado_multiplicacion*INVTSj
                                
                                #####HASTA ACÁ LLEGA EL PRIMER TÉRMINO, ARRANCA EL SEGUNDO. SE TOMAN VARIAS VARIABLES YA DEFINIDAS###

                                segundo_termino = sum(productos) * OPTSj
                                
                                #HASTA ACÁ LLEGA EL SEGUNDO TÉRMINO, ARRANCA EL TERCERO. SE TOMAN VARIAS VARIABLES YA DEFINIDAS###
                                Bjk = (np.tile(vector_resultante, (cantidad_de_ks, 1)).T) * (1- RECj*PS_en_ET)
                                matriz_transpuesta_cik = list(map(list, zip(*Cik)))
                                D = ((np.array(Yjk) * np.array(Bjk)) +  (np.array(Zik) * np.array(matriz_transpuesta_cik)))
                                E = generar_matriz_nueva(numeric_arrays[2])
                                if(evaluarCuartaRestriccion(D,E)):
                                    A = np.array(Yjk)
                                    B = np.array(Bjk)
                                    C = np.array(Ik)
                                    # Verificar si las matrices tienen la misma forma
                                    if A.shape != B.shape:
                                        print("Las matrices no tienen la misma forma.")
                                    # else:
                                    #     productos = productoEscalar2Matrices(A,B)
                                    #     resultado_final = sum(productos)
                                        # Convertir la lista de productos a un arreglo NumPy
                                    productos = productoEscalar2Matrices(A,B)
                                    vector_resultante = np.array(productos)
                                    resultado_multiplicacion = np.dot(vector_resultante, C)
                                    tercer_termino = resultado_multiplicacion*INVLk
                                    
                                    ######HASTA ACA TERCER TERMINO############

                                    matriz_transpuesta = list(map(list, zip(*Cik)))
                                    A = np.array(Zik)
                                    B = np.array(matriz_transpuesta)
                                    C = np.array(Ik)

                                    # Verificar si las matrices tienen la misma forma
                                    if A.shape != B.shape:
                                        print("Las matrices no tienen la misma forma.")
                                    # else:
                                    #     productos = productoEscalar2Matrices(A,B)
                                    #     resultado_final = sum(productos)
                                        # Convertir la lista de productos a un arreglo NumPy
                                    productos = productoEscalar2Matrices(A,B)
                                    vector_resultante = np.array(productos)
                                    resultado_multiplicacion = np.dot(vector_resultante, C)
                                    cuarto_termino = resultado_multiplicacion*INVLk
                                    
                                                       
                                    

                                    #################################HASTA ACÁ CUARTO TÉRMINO#############################

                                    A = np.array(Yjk)
                                    B = np.array(Bjk)

                                    # Verificar si las matrices tienen la misma forma
                                    if A.shape != B.shape:
                                        print("Las matrices no tienen la misma forma.")
                                    
                                    quinto_termino = calcular_termino_producto_escalar2(A,B,OPLk)

                                    ######################HASTA ACÁ QUINTO TÉRMINO###############################
                                    matriz_transpuesta = list(map(list, zip(*Cik)))
                                    A = np.array(Zik)
                                    B = np.array(matriz_transpuesta)
                                    # Verificar si las matrices tienen la misma forma
                                    if A.shape != B.shape:
                                        print("Las matrices no tienen la misma forma.")
                                    
                                    sexto_termino = calcular_termino_producto_escalar2(A,B,OPLk)

                                    ###############################HASTA ACÁ SEXTO TÉRMINO######################

                                    matriz_transpuesta = list(map(list, zip(*Aij)))
                                    A = np.array(Xij)
                                    B = np.array(matriz_transpuesta)
                                    C = np.array(Dij)

                                    # Verificar si las matrices tienen la misma forma
                                    if A.shape != B.shape:
                                        print("Las matrices no tienen la misma forma.")

                                    elif A.shape != C.shape:
                                        print("Las matrices no tienen la misma forma")
                                        
                                    septimo_termino = calcular_termino_producto_escalar3(A,B,C,CC)

                                    ###############################HASTA ACÁ SÉPTIMO TÉRMINO######################
                                    matriz_transpuesta = list(map(list, zip(*Cik)))
                                    A = np.array(Zik)
                                    B = np.array(matriz_transpuesta)
                                    C = np.array(Dik)
                                        # Verificar si las matrices tienen la misma forma
                                    if A.shape != B.shape:
                                        print("Las matrices no tienen la misma forma.")

                                    elif A.shape != C.shape:
                                        print("Las matrices no tienen la misma forma")
                                        
                                    # else:
                                    #     productos = productos = productoEscalar3Matrices(A,B,C)
                                    #     resultado_final = sum(productos)
                                    octavo_termino = calcular_termino_producto_escalar3(A,B,C,CC)

                                    ###############################HASTA ACÁ Octqavo TÉRMINO######################
                                    matriz_transpuesta = list(map(list, zip(*Bjk)))

                                    A = np.array(Yjk)
                                    B = np.array(Bjk)
                                    C = np.array(Djk)
                                        # Verificar si las matrices tienen la misma forma
                                    if A.shape != B.shape:
                                        print("Las matrices no tienen la misma forma.")

                                    elif A.shape != C.shape:
                                        print("Las matrices no tienen la misma forma")
                                
                                    noveno_termino = calcular_termino_producto_escalar3(A,B,C,TC)

                                    ###############################HASTA ACÁ noveno TÉRMINO######################

                                    # Mostrar el vector resultante calculado anteriormente, de hacer Xij(A) * Aij(B)
                                    A = np.array(Xij)
                                    B = np.array(list(map(list, zip(*Aij))))

                                    decimo_termino = calcular_termino_producto_escalar2(A,B,(RECj * SPj * PS_en_ET))
                                    
                                    ###############################HASTA ACÁ decimo TÉRMINO######################
                                    
                                    # Mostrar el vector resultante calculado anteriormente, de hacer Xij(A) * Aij(B)
                                    A = np.array(Zik)
                                    B = np.array(list(map(list, zip(*Cik))))

                                    
                                    decimoprimer_termino = calcular_termino_producto_escalar2(A,B,(RECk * SPk * PS_en_CA))
                                    ###############################HASTA ACÁ decimoprimer TÉRMINO######################
                                    sumatotal = (primer_termino + segundo_termino + tercer_termino + cuarto_termino + quinto_termino  + sexto_termino + septimo_termino + octavo_termino + noveno_termino - decimo_termino - decimoprimer_termino)
                                    if(sumatotal < valor_suma_minimo):
                                        valor_suma_minimo = sumatotal
                                        arreglo_terminos_optimos = [primer_termino[0], segundo_termino, tercer_termino[0], cuarto_termino[0], quinto_termino, sexto_termino, septimo_termino, octavo_termino, noveno_termino, decimo_termino, decimoprimer_termino]
                                        matriz_optimaXij = Xij
                                        matriz_optimaZik = Zik
                                        matriz_optimaYjk = Yjk
                            
print("Optima Xij", matriz_optimaXij)
print("Optima Zik", matriz_optimaZik)
print("Optima Yjk", matriz_optimaYjk)
print("suma valor minimo" ,valor_suma_minimo)



############################################################################################################






###################################### Comienza la segunda expresión ################

print("Comienzo de la segunda expresión  - Cálculo de camiones")

#Primer término de camiones

Qwc = Vc*Ec*Dc / (1+Rc)
TRcij= (T1c-(T2c+T3c))/((2*Dij/SPDc)+T4c+T5c)

matriz_transpuesta = list(map(list, zip(*Aij)))

A = np.array(matriz_transpuesta)
TRcij = np.array(TRcij)
B = np.round(TRcij)
TRcij = B
NCij = A / (Qwc*B)
A_sin_inf = np.nan_to_num(np.array(NCij), nan=0, posinf=0, neginf=0)
NCij = A_sin_inf


A = np.array(matriz_optimaXij)
NCij = np.array(NCij)
B = np.ceil(NCij)
NCij=B


# n_columnas = A.shape[1]
# productos = []
# productos = np.sum(A * B, axis=0)
productos = productoEscalar2Matrices(A,B)
primer_termino_segundaexp = np.sum(productos)


print("Resultado del primer termino segunda expresión:", primer_termino_segundaexp)

###################################### Hasta acá primer termino segunda exp ################


#Segundo término de camiones

Qwt = Vt*Et*Dt / (1+Rt)
TRtjk= (T1t-(T2t+T3t))/((2*Djk/SPDt)+T4t+T5t)

Bjk = generarNuevoBjK((np.array(matriz_optimaXij) * np.array(matriz_transpuesta)) * (1 - RECj*PS_en_ET),matriz_optimaYjk)
A = np.array(Bjk)
TRtjk = np.array(TRtjk)
# Redondear los valores de la matriz A hacia arriba al número entero más cercano
B = np.ceil(TRtjk)
TRtjk=B

NTjk = A / (Qwt*B)
A_sin_inf = np.nan_to_num(np.array(NTjk), nan=0, posinf=0, neginf=0)
NTjk = A_sin_inf


A = np.array(matriz_optimaYjk)
NTjk = np.array(NTjk)
B = np.ceil(NTjk)
NTjk=B
productos = productoEscalar2Matrices(A,B)
resultado_final = sum(productos)
segundo_termino_segundaexp = (resultado_final)

print("Resultado del segundo termino segunda expresión:", segundo_termino_segundaexp)


########################HASTA ACÁ SEGUNDO TÉRMINO DE LA SEGUNDA EXPRESIÓN #################

#Tercer término de camiones
TRcik= (T1c-(T2c+T3c))/((2*Dik/SPDc)+T4c+T5c)
matriz_transpuesta = list(map(list, zip(*Cik)))


A = np.array(matriz_transpuesta)
TRcik = np.array(TRcik)


B = np.ceil(TRcik)
TRcik = B



NCik = A / (Qwc*B)
A_sin_inf = np.nan_to_num(np.array(NCik), nan=0, posinf=0, neginf=0)
NCik = A_sin_inf



A = np.array(matriz_optimaZik)
NCik = np.array(NCik)
B = np.ceil(NCik)
NCik=B
productos = productoEscalar2Matrices(A,B)
resultado_final = sum(productos)
tercer_termino_segundaexp = (resultado_final)


#########################SUMA TOTAL DE CAMIONES

print(" La flota total de camiones se constituye por ", primer_termino_segundaexp + tercer_termino_segundaexp , "de camiones recolectores y ", segundo_termino_segundaexp, "de camiones transportadores, dando un total de " , primer_termino_segundaexp + segundo_termino_segundaexp + tercer_termino_segundaexp, "camiones")


print("Comienzo de la tercer expresión  - Cálculo de gases de EI en kg de CO2 eq")


A = np.array(matriz_optimaXij)
NCij = np.array(NCij)
B = np.ceil(NCij)
NCij=B
B = A*B

A = np.array(matriz_optimaXij)
B = np.array(NCij)
C = np.array(Dij)
D = np.array(TRcij)

    # Verificar si las matrices tienen la misma forma
if A.shape != B.shape:
    print("Las matrices no tienen la misma forma.")

elif A.shape != C.shape:
    print("Las matrices no tienen la misma forma")
    
elif A.shape != D.shape:
    print("Las matrices no tienen la misma forma")
    
else: 
    productos = productoEscalar4Matrices(A,B,C,D)
    # Sumar los productos escalares individuales para obtener el resultado final
    resultado_final = sum(productos)

    
    primer_termino_tercerexp = resultado_final*(Uc/KPLc)*2
    
    print("Resultado del primer término de la tercer expresión:", primer_termino_tercerexp)
    ###################################### Hasta acá primer termino de la tercer expresión ################
    

A = np.array(matriz_optimaYjk)
B = np.array(NTjk)
C = np.array(Djk)
D = np.array(TRtjk)


# print("AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA")
# print(matriz_optimaYjk)
# print(NTjk)
# print(Djk)
# print(TRtjk)

    # Verificar si las matrices tienen la misma forma
if A.shape != B.shape:
    print("Las matrices no tienen la misma forma.")

elif A.shape != C.shape:
    print("Las matrices no tienen la misma forma")
    
elif A.shape != D.shape:
    print("Las matrices no tienen la misma forma")
    
else: 
    n_columnas = A.shape[1]
    productos = []
    
    # Calcular el producto escalar entre las columnas correspondientes
    productos = productoEscalar4Matrices(A,B,C,D)
    # Sumar los productos escalares individuales para obtener el resultado final
    resultado_final = sum(productos)
    #print("\nResultado final del producto escalar entre las columnas:", resultado_final)

    segundo_termino_tercerexp = resultado_final*(Ut/KPLt)*2
    
    
###################################### Hasta acá segundo termino de la tercer expresión ################  

A = np.array(matriz_optimaZik)
B = np.array(NCik)
C = np.array(Dik)
D = np.array(TRcik)

    # Verificar si las matrices tienen la misma forma
if A.shape != B.shape:
    print("Las matrices no tienen la misma forma.")

elif A.shape != C.shape:
    print("Las matrices no tienen la misma forma")
    
elif A.shape != D.shape:
    print("Las matrices no tienen la misma forma")
    
else: 
    
    
    # Calcular el producto escalar entre las columnas correspondientes
    productos = productoEscalar4Matrices(A,B,C,D)
    # Sumar los productos escalares individuales para obtener el resultado final
    resultado_final = sum(productos)
    #print("\nResultado final del producto escalar entre las columnas:", resultado_final)

    tercer_termino_tercerexp = resultado_final*(Ut/KPLt)*2
    
    

    
    
#################################################################################

#Se borra la solapa "Resultados"
# Nombre del archivo Excel
# escenario = hoja_INPUTS_Generales['A2'].value
archivo_federico = openpyxl.load_workbook('modelo.xlsm')
# Seleccionar la hoja del archivo "Federico"
hoja_federico = archivo_federico['INPUTS_Generales']
# Obtener el valor de la celda A2 en el archivo "Federico"
escenario = hoja_federico['A2'].value

nombre_archivo = f"RESULTADOS_{escenario}.xlsx"
# Cargar el archivo Excel existente (xlsm)

# datos_excel
nuevo_libro = openpyxl.Workbook()


# Seleccionar la hoja en la que se imprimirá el texto
nueva_hoja = nuevo_libro.create_sheet(title='Resultados')

column_widths = {
    'A': 45,
    'B': 15,
    'C': 45,
    'D': 15,
    'E': 45,
    'F': 15,
}

for column, width in column_widths.items():
    nueva_hoja.column_dimensions[column].width = width



for t in arreglo_terminos_optimos:
    print ("termino: ", t)
    

# Eliminar corchetes y formatear el número con dos decimales
valor_formateado = round(valor_suma_minimo, 2)
#valor_formateado = "{:.2f}".format(valor_suma_minimo[0])
nueva_hoja['A4'] = "Costo total del sistema [Usd]"
nueva_hoja['B4'] = valor_formateado

# Eliminar corchetes y formatear el número con dos decimales
valor_formateado = "{:.2f}".format(arreglo_terminos_optimos[0])
nueva_hoja['A5'] = "Costo de Inversión ET"
nueva_hoja['B5'] = valor_formateado

valor_formateado = round(arreglo_terminos_optimos[1], 1)
nueva_hoja['A6'] = "Costo Operativo ET"
nueva_hoja['B6'] = valor_formateado

# Eliminar corchetes y formatear el número con dos decimales
valor_formateado1 = "{:.2f}".format(arreglo_terminos_optimos[2])
valor_formateado2 = "{:.2f}".format(arreglo_terminos_optimos[3])
# Convertir los valores formateados a números flotantes y redondearlos
valor_redondeado1 = round(float(valor_formateado1), 1)
valor_redondeado2 = round(float(valor_formateado2), 1)
# Asignar la suma de los valores redondeados a la celda B7
nueva_hoja['A7'] = "Costo de Inversión CA"
nueva_hoja['B7'] = valor_redondeado1 + valor_redondeado2


valor_formateado1 = round(arreglo_terminos_optimos[4], 1)
valor_formateado2 = round(arreglo_terminos_optimos[5], 1)
nueva_hoja['A8'] = "Costo Operativo CA"
nueva_hoja['B8'] = valor_formateado1 + valor_formateado2

valor_formateado1 = round(arreglo_terminos_optimos[6], 1)
valor_formateado2 = round(arreglo_terminos_optimos[7], 1)
nueva_hoja['A9'] = "Costo de Transporte Compactadores"
nueva_hoja['B9'] = valor_formateado1 + valor_formateado2

valor_formateado = round(arreglo_terminos_optimos[8], 1)
nueva_hoja['A10'] = "Costo de Transporte de Transferencia"
nueva_hoja['B10'] = valor_formateado

valor_formateado = round(arreglo_terminos_optimos[9], 1)
nueva_hoja['A11'] = "Costo de Venta de material recuperado en ET"
nueva_hoja['B11'] = valor_formateado

valor_formateado = round(arreglo_terminos_optimos[10], 1)
nueva_hoja['A12'] = "Costo de Venta de Material recuperado en CA"
nueva_hoja['B12'] = valor_formateado


#Comienza a imprimirse camiones
nueva_hoja['C4'] = "Cantidad de camiones totales"
nueva_hoja['D4'] = (primer_termino_segundaexp + segundo_termino_segundaexp + tercer_termino_segundaexp)

nueva_hoja['C5'] = "Compactadores a ET [U]"
nueva_hoja['D5'] = (primer_termino_segundaexp)

nueva_hoja['C6'] = "Compactadores a CA"
nueva_hoja['D6'] = (tercer_termino_segundaexp)

nueva_hoja['C7'] = "De Transferencia"
nueva_hoja['D7'] = (segundo_termino_segundaexp)


#Comienza a imprimirse gases  GEI
nueva_hoja['E4'] = "Gases GEI totales [kgCO2eq]"
nueva_hoja['F4'] = round((primer_termino_tercerexp + segundo_termino_tercerexp + tercer_termino_tercerexp),2)

nueva_hoja['E5'] = "Gases GEI compactadores (a ET)"
nueva_hoja['F5'] = round((primer_termino_tercerexp),2)

nueva_hoja['E6'] = "Gases GEI compactadores (a CA)"
nueva_hoja['F6'] = round((tercer_termino_tercerexp),2)

nueva_hoja['E7'] = "Gases GEI transferencia"
nueva_hoja['F7'] = round((segundo_termino_tercerexp),2)




if texto_i2 == "-":
    frase_titulo = "REGIÓN : " + texto_region + " - Conformada por " + texto_i1
elif texto_i3 == "-":
    frase_titulo = "REGIÓN : " + texto_region + " - Conformada por " + texto_i1 + " y " + texto_i2
elif texto_i4 == "-":
    frase_titulo = "REGIÓN : " + texto_region + " - Conformada por " + texto_i1 + ", " + texto_i2 + " y " + texto_i3
elif texto_i5 == "-":
    frase_titulo = "REGIÓN : " + texto_region + " - Conformada por " + texto_i1 + ", " + texto_i2 + ", " + texto_i3 + " y " + texto_i4
elif texto_i6 == "-":
    frase_titulo = "REGIÓN : " + texto_region + " - Conformada por " + texto_i1 + ", " + texto_i2 + ", " + texto_i3 + ", " + texto_i4 + " y " + texto_i5
elif texto_i7 == "-":
    frase_titulo = "REGIÓN : " + texto_region + " - Conformada por " + texto_i1 + ", " + texto_i2 + ", " + texto_i3 + ", " + texto_i4 + ", " + texto_i5 + " y " + texto_i6
else: 
    frase_titulo = "REGIÓN : " + texto_region + " - Conformada por " + texto_i1

        
nueva_hoja['A1'] = str(frase_titulo)

#Comienza a imprimirse las matrices
nueva_hoja['A14'] = "Distribución de residuos del sistema [t/d]"
# Obtener la celda y aplicar el formato de negrita
celda = nueva_hoja['A14']
celda.font = Font(bold=True)

# Encontrar la última fila en la hoja
ultima_fila = nueva_hoja.max_row     
# Matriz de datos
matriz_transpuesta=list(map(list, zip(*Aij)))

datos_resultantes_A = (np.array(matriz_transpuesta) * np.array(matriz_optimaXij))
    
    
# Convertir cada elemento de la matriz a texto
datos_resultantes_A_texto = [
    [str(elemento) for elemento in fila] for fila in datos_resultantes_A
]

# Encabezados para las filas (I1, I2, I3, ...)
encabezados_filas = [f"I{i}" for i in range(1, len(datos_resultantes_A_texto) + 1)]

# Encabezados para las columnas (J1, J2, J3, ...)
encabezados_columnas = [f"J{j}" for j in range(1, len(datos_resultantes_A_texto[0]) + 1)]

# Insertar encabezados de filas y columnas en la matriz de datos
datos_con_encabezados = [
    ["Sistema Loc-ET"] + encabezados_columnas
] + [
    [encabezado_fila] + fila for encabezado_fila, fila in zip(encabezados_filas, datos_resultantes_A_texto)
]

# Obtener la suma de cada columna
sumas_columnas = [sum(columna) for columna in zip(*datos_resultantes_A)]

# Agregar la fila con las sumas al final de la matriz
fila_suma = ['SUMA'] + sumas_columnas
datos_con_encabezados.append(fila_suma)

# Escribir la matriz con encabezados y suma en la hoja del Excel
for fila, fila_datos in enumerate(datos_con_encabezados, start=1):
    for columna, valor in enumerate(fila_datos, start=1):
        nueva_hoja.cell(row=ultima_fila + 1 + fila, column=columna, value=str(valor))  # Convertir el valor a texto


#Imprimo segunda matriz debajo de la anterior"
# Encontrar la última fila en la hoja
ultima_fila = nueva_hoja.max_row 

# Matriz B
datos_resultantes_b = Bjk

    
# Convertir cada elemento de la matriz a texto
datos_resultantes_b_texto = [
    [str(elemento) for elemento in fila] for fila in datos_resultantes_b
]

# Encabezados para las filas (I1, I2, I3, ...)
encabezados_filas = [f"J{i}" for i in range(1, len(datos_resultantes_b_texto) + 1)]

# Encabezados para las columnas (J1, J2, J3, ...)
encabezados_columnas = [f"K{k}" for k in range(1, len(datos_resultantes_b_texto[0]) + 1)]

# Insertar encabezados de filas y columnas en la matriz de datos
datos_con_encabezados = [
    ["Sistema ET-CA"] + encabezados_columnas
] + [
    [encabezado_fila] + fila for encabezado_fila, fila in zip(encabezados_filas, datos_resultantes_b_texto)
]

# Obtener la suma de cada columna
sumas_columnas = [sum(columna) for columna in zip(*datos_resultantes_b)]

# Agregar la fila con las sumas al final de la matriz
fila_suma = ['SUMA'] + sumas_columnas
datos_con_encabezados.append(fila_suma)

# Escribir la matriz con encabezados y suma en la hoja del Excel
for fila, fila_datos in enumerate(datos_con_encabezados, start=1):
    for columna, valor in enumerate(fila_datos, start=1):
        nueva_hoja.cell(row=ultima_fila + 1 + fila, column=columna, value=str(valor))  # Convertir el valor a texto





# Encontrar la última fila en la hoja
ultima_fila = nueva_hoja.max_row     
# Matriz de datos
matriz_transpuesta=list(map(list, zip(*Cik)))

datos_resultantes_c = (np.array(matriz_transpuesta) * np.array(matriz_optimaZik))

    
# Convertir cada elemento de la matriz a texto
datos_resultantes_c_texto = [
    [str(elemento) for elemento in fila] for fila in datos_resultantes_c
]

# Encabezados para las filas (I1, I2, I3, ...)
encabezados_filas = [f"I{i}" for i in range(1, len(datos_resultantes_c_texto) + 1)]

# Encabezados para las columnas (J1, J2, J3, ...)
encabezados_columnas = [f"K{k}" for k in range(1, len(datos_resultantes_c_texto[0]) + 1)]

# Insertar encabezados de filas y columnas en la matriz de datos
datos_con_encabezados = [
    ["Sistema ET-CA"] + encabezados_columnas
] + [
    [encabezado_fila] + fila for encabezado_fila, fila in zip(encabezados_filas, datos_resultantes_c_texto)
]

# Obtener la suma de cada columna
sumas_columnas = [sum(columna) for columna in zip(*datos_resultantes_c)]

# Agregar la fila con las sumas al final de la matriz
fila_suma = ['SUMA'] + sumas_columnas
datos_con_encabezados.append(fila_suma)

# Escribir la matriz con encabezados y suma en la hoja del Excel
for fila, fila_datos in enumerate(datos_con_encabezados, start=1):
    for columna, valor in enumerate(fila_datos, start=1):
        nueva_hoja.cell(row=ultima_fila + 1 + fila, column=columna, value=str(valor))  # Convertir el valor a texto





# Encontrar la última fila con texto en la hoja de Excel
ultima_fila_con_texto = max((i for i, row in enumerate(nueva_hoja.iter_rows(), start=1) if any(cell.value for cell in row)), default=0)

# Calcula la posición para escribir tu información (2 filas por debajo de la última con texto)
ultima_fila = ultima_fila_con_texto + 2  # Dos filas de espacio

# Texto a imprimir dos filas por debajo de la última fila con información
texto_a_imprimir = 'Distribución de camiones en todo el sistema [u]'


# Escribir el texto en la celda A (columna 1) dos filas por debajo de la última fila con información
celda = nueva_hoja.cell(row=ultima_fila + 1, column=1, value=texto_a_imprimir)

# Aplicar el formato de negrita a la celda
celda.font = Font(bold=True)


# Encontrar la última fila en la hoja
ultima_fila = nueva_hoja.max_row     
# Matriz de datos


datos_resultantes_A = (np.array(NCij))*(np.array(matriz_optimaXij))
    
    
# Convertir cada elemento de la matriz a texto
datos_resultantes_A_texto = [
    [str(elemento) for elemento in fila] for fila in datos_resultantes_A
]

# Encabezados para las filas (I1, I2, I3, ...)
encabezados_filas = [f"I{i}" for i in range(1, len(datos_resultantes_A_texto) + 1)]

# Encabezados para las columnas (J1, J2, J3, ...)
encabezados_columnas = [f"J{j}" for j in range(1, len(datos_resultantes_A_texto[0]) + 1)]

# Insertar encabezados de filas y columnas en la matriz de datos
datos_con_encabezados = [
    ["Sist. Loc-ET"] + encabezados_columnas
] + [
    [encabezado_fila] + fila for encabezado_fila, fila in zip(encabezados_filas, datos_resultantes_A_texto)
]

# Obtener la suma de cada columna
sumas_columnas = [sum(columna) for columna in zip(*datos_resultantes_A)]

# Agregar la fila con las sumas al final de la matriz
fila_suma = ['SUMA'] + sumas_columnas
datos_con_encabezados.append(fila_suma)

# Escribir la matriz con encabezados y suma en la hoja del Excel
for fila, fila_datos in enumerate(datos_con_encabezados, start=1):
    for columna, valor in enumerate(fila_datos, start=1):
        nueva_hoja.cell(row=ultima_fila + 1 + fila, column=columna, value=str(valor))  # Convertir el valor a texto
        
        
        
# Encontrar la última fila en la hoja
ultima_fila = nueva_hoja.max_row     
# Matriz de datos


datos_resultantes_A = (np.array(NTjk))*(np.array(matriz_optimaYjk))
    
    
# Convertir cada elemento de la matriz a texto
datos_resultantes_A_texto = [
    [str(elemento) for elemento in fila] for fila in datos_resultantes_A
]

# Encabezados para las filas (I1, I2, I3, ...)
encabezados_filas = [f"J{j}" for j in range(1, len(datos_resultantes_A_texto) + 1)]

# Encabezados para las columnas (J1, J2, J3, ...)
encabezados_columnas = [f"K{k}" for k in range(1, len(datos_resultantes_A_texto[0]) + 1)]

# Insertar encabezados de filas y columnas en la matriz de datos
datos_con_encabezados = [
    ["Sistema ET-CA"] + encabezados_columnas
] + [
    [encabezado_fila] + fila for encabezado_fila, fila in zip(encabezados_filas, datos_resultantes_A_texto)
]

# Obtener la suma de cada columna
sumas_columnas = [sum(columna) for columna in zip(*datos_resultantes_A)]

# Agregar la fila con las sumas al final de la matriz
fila_suma = ['SUMA'] + sumas_columnas
datos_con_encabezados.append(fila_suma)

# Escribir la matriz con encabezados y suma en la hoja del Excel
for fila, fila_datos in enumerate(datos_con_encabezados, start=1):
    for columna, valor in enumerate(fila_datos, start=1):
        nueva_hoja.cell(row=ultima_fila + 1 + fila, column=columna, value=str(valor))  # Convertir el valor a texto    
        
        
        
 # Encontrar la última fila en la hoja
ultima_fila = nueva_hoja.max_row     
# Matriz de datos


datos_resultantes_A = (np.array(NCik))*(np.array(matriz_optimaZik))
    
    
# Convertir cada elemento de la matriz a texto
datos_resultantes_A_texto = [
    [str(elemento) for elemento in fila] for fila in datos_resultantes_A
]

# Encabezados para las filas (I1, I2, I3, ...)
encabezados_filas = [f"I{i}" for i in range(1, len(datos_resultantes_A_texto) + 1)]

# Encabezados para las columnas (J1, J2, J3, ...)
encabezados_columnas = [f"K{k}" for k in range(1, len(datos_resultantes_A_texto[0]) + 1)]

# Insertar encabezados de filas y columnas en la matriz de datos
datos_con_encabezados = [
    ["Sistema Loc-CA"] + encabezados_columnas
] + [
    [encabezado_fila] + fila for encabezado_fila, fila in zip(encabezados_filas, datos_resultantes_A_texto)
]

# Obtener la suma de cada columna
sumas_columnas = [sum(columna) for columna in zip(*datos_resultantes_A)]

# Agregar la fila con las sumas al final de la matriz
fila_suma = ['SUMA'] + sumas_columnas
datos_con_encabezados.append(fila_suma)

# Escribir la matriz con encabezados y suma en la hoja del Excel
for fila, fila_datos in enumerate(datos_con_encabezados, start=1):
    for columna, valor in enumerate(fila_datos, start=1):
        nueva_hoja.cell(row=ultima_fila + 1 + fila, column=columna, value=str(valor))  # Convertir el valor a texto    
        
        
        
        
        

# Encontrar la última fila con texto en las columnas de A a F
last_row = nueva_hoja.max_row

# Aplicar borde grueso a las celdas del rango A3:F{last_row}
for row in range(3, last_row + 1):
    for column in ['A', 'F']:
        cell = nueva_hoja[f'{column}{row}']
        if column == 'A':
            cell.border = Border(left=Side(style='thick'))
        elif column == 'F':
            cell.border = Border(right=Side(style='thick'))

# Aplicar borde grueso a las celdas de la fila A3:F{last_row}
for column in ['B', 'C', 'D', 'E']:
    cell_top = nueva_hoja[f'{column}3']
    cell_bottom = nueva_hoja[f'{column}{last_row}']
    
    cell_top.border = Border(top=Side(style='thick'))
    cell_bottom.border = Border(bottom=Side(style='thick'))

# Aplicar borde inferior en A(maxrow) y F(maxrow)
cell_A_bottom = nueva_hoja[f'A{last_row}']
cell_F_bottom = nueva_hoja[f'F{last_row}']

cell_A_bottom.border = Border(left=Side(style='thick'), bottom=Side(style='thick'))
cell_F_bottom.border = Border(right=Side(style='thick'), bottom=Side(style='thick'))

# Aplicar estilo a las celdas de la columna A a la F y de la fila 1 a la 100
for row in range(1, 101):
    for column in ['A', 'B', 'C', 'D', 'E', 'F']:
        cell = nueva_hoja[f'{column}{row}']
        cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

# Combina las celdas desde A1 hasta F2
nueva_hoja.merge_cells('A1:F2')

# Aplicar estilo a las celdas combinadas
for row in nueva_hoja.iter_rows(min_row=1, max_row=2, min_col=1, max_col=6):
    for cell in row:
        cell.fill = PatternFill(start_color="95B3D7", end_color="95B3D7", fill_type="solid")
        cell.font = Font(bold=True, size=16)
        cell.border = Border(left=Side(style='thick'),
                             right=Side(style='thick'),
                             top=Side(style='thick'),
                             bottom=Side(style='thick'))
              

#Iterar sobre todas las ceeldas y establecer alineación
for row in nueva_hoja.iter_rows():
    for cell in row:
        cell.alignment = Alignment (horizontal='center', vertical='center')
        
        
# Reemplazar I1,2,3 por nombres cargados

# Define un diccionario que mapee los valores "I1", "I2", etc., con los textos deseados
mapeo_textos = {
    "I1": texto_i1,
    "I2": texto_i2,
    "I3": texto_i3,
    "I4": texto_i4,
    "I5": texto_i5,
    "I6": texto_i6
}

# Itera sobre todas las celdas de la hoja
for fila in nueva_hoja.iter_rows():
    for celda in fila:
        # Verifica si el valor de la celda está en el diccionario de mapeo
        if celda.value in mapeo_textos:
            # Reemplaza el valor de la celda con el texto correspondiente del diccionario
            celda.value = mapeo_textos[celda.value]


# Define un diccionario que mapee los valores "I1", "I2", etc., con los textos deseados
mapeo_textos = {
    "J1": "ET " + texto_i1,
    "J2": "ET " + texto_i2,
    "J3": "ET " + texto_i3,
    "J4": "ET " + texto_i4,
    "J5": "ET " + texto_i5,
    "J6": "ET " + texto_i6
}

# Itera sobre todas las celdas de la hoja
for fila in nueva_hoja.iter_rows():
    for celda in fila:
        # Verifica si el valor de la celda está en el diccionario de mapeo
        if celda.value in mapeo_textos:
            # Reemplaza el valor de la celda con el texto correspondiente del diccionario
            celda.value = mapeo_textos[celda.value]
            
            
            
# Itera sobre todas las celdas de la hoja
for fila in nueva_hoja.iter_rows():
    for celda in fila:
        # Verifica si el valor de la celda es "I1"
        if celda.value == "K1":
            # Reemplaza "I1" por el texto deseado
            celda.value = "CA Regional"

# Verifica si la hoja "Sheet" está presente en el archivo
if 'Sheet' in nuevo_libro.sheetnames:
    # Elimina la hoja "Sheet"
    nuevo_libro.remove(nuevo_libro['Sheet'])
    
    
# Iterar sobre todas las celdas en la hoja de cálculo
for fila in nueva_hoja.iter_rows():
    for celda in fila:
        # Verificar si la celda contiene texto y si contiene un punto
        if isinstance(celda.value, str) and "." in celda.value:
            # Reemplazar los puntos por comas
            celda.value = celda.value.replace(".", ",")
    
    

# Guardar los cambios en el archivo Excel (xlsm)
nuevo_libro.save(nombre_archivo)


final =  datetime.now() - hora_actual
print("tiempo total de ejecución en minutos: ", final.total_seconds()/60)




